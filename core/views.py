import os
import csv
import calendar as cal_module
from datetime import datetime, date, timedelta
from functools import wraps
from io import BytesIO, StringIO

import uuid

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.db.models import Sum, Avg, Q, Count
from django.contrib import messages
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.core.exceptions import PermissionDenied
from django.views.decorators.http import require_POST
from django.core.management import call_command
from django.core.paginator import Paginator
from django.utils import timezone
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
import re

from .models import (Task, Note, Notification, Tag, Attachment,
                     Department, Profile, TaskTemplate, TaskComment, TaskLog, LoginLog,
                     PasswordHistory, AdminLog, TaskWatcher)
from .forms import (TaskCreateForm, TaskUpdateForm, NoteForm,
                    AdminUserCreateForm, AdminUserEditForm, AdminDepartmentForm,
                    CommentForm, TaskTemplateForm, ProfileForm)

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors as rl_colors
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ─── Helpers ──────────────────────────────────────────────────────

def crear_notificacion(usuario, tarea, mensaje):
    Notification.objects.create(usuario=usuario, tarea=tarea, mensaje=mensaje)
    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        f'notifications_{usuario.id}',
        {
            'type': 'notification_event',
            'count': Notification.objects.filter(usuario=usuario, leido=False).count(),
            'message': mensaje,
        },
    )


def crear_log(tarea, usuario, accion, detalle=''):
    TaskLog.objects.create(tarea=tarea, usuario=usuario, accion=accion, detalle=detalle)


def tareas_visibles(usuario):
    profile = Profile.objects.filter(usuario=usuario).first()
    if not profile:
        return Task.objects.filter(usuario=usuario)
    deptos = profile.departamentos.all()
    if profile.rol in ('admin', 'moderador'):
        qs = Task.objects.all()
        if deptos:
            qs = qs.filter(Q(usuario=usuario) | Q(departamento__in=deptos))
        return qs.distinct()
    qs = Task.objects.filter(usuario=usuario)
    if deptos:
        qs = qs | Task.objects.filter(departamento__in=deptos)
    return qs.distinct()


def _admin_log(usuario, accion, detalle=''):
    AdminLog.objects.create(usuario=usuario, accion=accion, detalle=detalle)


# ─── Role decorators ──────────────────────────────────────────────

def admin_required(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        p = Profile.objects.filter(usuario=request.user).first()
        if not p or p.rol != 'admin':
            raise PermissionDenied
        return view_func(request, *args, **kwargs)
    return _wrapped


def admin_mod_required(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        p = Profile.objects.filter(usuario=request.user).first()
        if not p or p.rol not in ('admin', 'moderador'):
            raise PermissionDenied
        return view_func(request, *args, **kwargs)
    return _wrapped


# ═══════════════════════════════════════════════════════════════════
#  HOME
# ═══════════════════════════════════════════════════════════════════

@login_required
def index(request):
    hoy = date.today()
    qs = tareas_visibles(request.user)
    chart_estados = {
        'labels': ['Pendientes', 'En Proceso', 'Terminadas'],
        'data': [
            qs.filter(estado='pendiente').count(),
            qs.filter(estado='proceso').count(),
            qs.filter(estado='terminada').count(),
        ],
    }
    prioridad_counts = {
        'urgente': qs.filter(prioridad='urgente').count(),
        'alta': qs.filter(prioridad='alta').count(),
        'media': qs.filter(prioridad='media').count(),
        'baja': qs.filter(prioridad='baja').count(),
    }
    # Tendencia: completadas por día últimos 14 días
    trend_labels = []
    trend_data = []
    for i in range(13, -1, -1):
        d = hoy - timedelta(days=i)
        trend_labels.append(d.strftime('%d/%m'))
        trend_data.append(qs.filter(estado='terminada', fecha_completada=d).count())
    context = {
        'tareas_pendientes': qs.filter(estado='pendiente'),
        'tareas_proceso': qs.filter(estado='proceso'),
        'tareas_hoy': qs.filter(fecha_vencimiento=hoy),
        'tareas_urgentes': qs.filter(prioridad='urgente', estado__in=['pendiente', 'proceso']),
        'notas_recientes': Note.objects.filter(usuario=request.user)[:5],
        'notificaciones': Notification.objects.filter(usuario=request.user, leido=False),
        'total_pendientes': qs.filter(estado='pendiente').count(),
        'total_proceso': qs.filter(estado='proceso').count(),
        'total_terminadas': qs.filter(estado='terminada').count(),
        'etiquetas': Tag.objects.filter(usuario=request.user),
        'total_tareas': qs.count(),
        'eficiencia': _calcular_eficiencia(qs),
        'chart_estados': chart_estados,
        'chart_prioridades': prioridad_counts,
        'trend_labels': trend_labels,
        'trend_data': trend_data,
    }
    return render(request, 'core/index.html', context)


def _calcular_eficiencia(qs):
    completadas = qs.filter(estado='terminada')
    total = completadas.count()
    if total == 0:
        return 0
    con_horas = completadas.filter(horas_estimadas__isnull=False, horas_tomadas__isnull=False)
    if con_horas.count() == 0:
        return 0
    suma = 0
    for t in con_horas:
        if t.horas_estimadas > 0:
            suma += (t.horas_estimadas / t.horas_tomadas) * 100
    return round(suma / con_horas.count(), 1)


# ═══════════════════════════════════════════════════════════════════
#  TASKS
# ═══════════════════════════════════════════════════════════════════

class TaskListView(LoginRequiredMixin, ListView):
    model = Task
    template_name = 'core/task_list.html'
    context_object_name = 'tareas'
    paginate_by = 30

    def get_queryset(self):
        qs = tareas_visibles(self.request.user)
        estado = self.request.GET.get('estado')
        prioridad = self.request.GET.get('prioridad')
        q = self.request.GET.get('q')
        etiqueta = self.request.GET.get('etiqueta')
        if estado: qs = qs.filter(estado=estado)
        if prioridad: qs = qs.filter(prioridad=prioridad)
        if q: qs = qs.filter(titulo__icontains=q)
        if etiqueta: qs = qs.filter(etiquetas__id=etiqueta)
        return qs.prefetch_related('subtareas').select_related('usuario', 'creado_por')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        for k in ('estado', 'prioridad', 'q', 'etiqueta'):
            ctx[f'{k}_actual'] = self.request.GET.get(k, '')
        ctx['etiquetas'] = Tag.objects.filter(usuario=self.request.user)
        ctx['user_es_admin'] = Profile.objects.filter(usuario=self.request.user, rol='admin').exists()
        # Subtask progress for each task on current page
        ctx['subtask_progress'] = {}
        for t in ctx['tareas']:
            subs = list(t.subtareas.all())
            if subs:
                total = len(subs)
                done = sum(1 for s in subs if s.estado == 'terminada')
                ctx['subtask_progress'][t.id] = {'total': total, 'done': done}
        return ctx


@login_required
@require_POST
def bulk_action(request):
    action = request.POST.get('action')
    task_ids = request.POST.getlist('task_ids')
    if not task_ids or not action:
        messages.error(request, 'Selecciona al menos una tarea.')
        return redirect('task_list')

    qs = tareas_visibles(request.user).filter(id__in=task_ids)
    es_admin = Profile.objects.filter(usuario=request.user, rol='admin').exists()

    if action == 'completar':
        count = qs.exclude(estado='terminada').update(estado='terminada',
            fecha_completada=datetime.now())
        messages.success(request, f'{count} tarea(s) marcada(s) como terminada(s).')
        for t in qs:
            TaskLog.objects.create(tarea=t, usuario=request.user, accion='completada_masiva')

    elif action == 'eliminar' and es_admin:
        count = qs.count()
        qs.delete()
        _admin_log(request.user, 'eliminar_tarea', f'{count} tarea(s) eliminada(s) masivamente')
        messages.success(request, f'{count} tarea(s) eliminada(s).')

    elif action == 'eliminar' and not es_admin:
        messages.error(request, 'Solo los administradores pueden eliminar tareas.')

    elif action == 'reasignar':
        nuevo_user_id = request.POST.get('nuevo_usuario')
        if nuevo_user_id and es_admin:
            try:
                nuevo_user = User.objects.get(pk=nuevo_user_id)
                count = qs.update(usuario=nuevo_user)
                _admin_log(request.user, 'reasignar_masiva', f'{count} tarea(s) reasignada(s) a {nuevo_user.username}')
                messages.success(request, f'{count} tarea(s) reasignada(s).')
                for t in qs:
                    TaskLog.objects.create(tarea=t, usuario=request.user,
                        accion=f'reasignada a {nuevo_user.username}')
            except User.DoesNotExist:
                messages.error(request, 'Usuario no válido.')

    return redirect('task_list')


class TaskDetailView(LoginRequiredMixin, DetailView):
    model = Task
    template_name = 'core/task_detail.html'
    context_object_name = 'tarea'

    def get_queryset(self):
        return tareas_visibles(self.request.user)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['subtareas'] = self.object.subtareas.all()
        ctx['archivos'] = self.object.archivos.all()
        ctx['puede_eliminar'] = Profile.objects.filter(usuario=self.request.user, rol='admin').exists()
        ctx['comentarios'] = self.object.task_comentarios.all()
        ctx['logs'] = self.object.logs.all()[:20]
        ctx['comment_form'] = CommentForm()
        ctx['is_watching'] = TaskWatcher.objects.filter(usuario=self.request.user, tarea=self.object).exists()
        ctx['task_locked'] = self.object.locked_by is not None and self.object.locked_by != self.request.user
        ctx['locked_by_user'] = self.object.locked_by.username if self.object.locked_by else None
        return ctx


@login_required
def add_comment(request, pk):
    tarea = get_object_or_404(Task, pk=pk)
    if tarea not in tareas_visibles(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        form = CommentForm(request.POST)
        if form.is_valid():
            contenido = form.cleaned_data['contenido']
            TaskComment.objects.create(
                tarea=tarea, usuario=request.user,
                contenido=contenido,
            )
            crear_log(tarea, request.user, 'editada', 'Nuevo comentario añadido')
            # Menciones @usuario
            for match in re.finditer(r'@(\w+)', contenido):
                try:
                    mencionado = User.objects.get(username=match.group(1))
                    if mencionado != request.user:
                        crear_notificacion(mencionado, tarea,
                            f'{request.user.username} te mencionó en "{tarea.titulo}"')
                except User.DoesNotExist:
                    pass
            # Notificar a observadores
            for watcher in TaskWatcher.objects.filter(tarea=tarea).exclude(usuario=request.user):
                crear_notificacion(watcher.usuario, tarea,
                    f'Nuevo comentario en "{tarea.titulo}" por {request.user.username}')
            messages.success(request, 'Comentario añadido.')
    return redirect('task_detail', pk=pk)


class TaskCreateView(LoginRequiredMixin, CreateView):
    model = Task
    form_class = TaskCreateForm
    template_name = 'core/task_form.html'

    def get_form_kwargs(self):
        kw = super().get_form_kwargs()
        kw['usuario_actual'] = self.request.user.pk
        kw['user'] = self.request.user
        return kw

    def form_valid(self, form):
        form.instance.usuario = form.cleaned_data['usuario']
        form.instance.creado_por = self.request.user
        form.instance.estado = 'pendiente'
        resp = super().form_valid(form)
        for f in self.request.FILES.getlist('archivos'):
            Attachment.objects.create(tarea=self.object, archivo=f, nombre=f.name)
        ets = form.cleaned_data.get('etiquetas')
        if ets: self.object.etiquetas.set(ets)
        crear_notificacion(self.object.usuario, self.object, f'Nueva tarea: "{self.object.titulo}"')
        # Notificar a observadores del padre si es subtarea
        if self.object.parent:
            for watcher in TaskWatcher.objects.filter(tarea=self.object.parent).exclude(usuario=self.request.user):
                crear_notificacion(watcher.usuario, self.object,
                    f'Nueva subtarea en "{self.object.parent.titulo}" por {self.request.user.username}')
        crear_log(self.object, self.request.user, 'creada', f'Creada por {self.request.user.username}')
        messages.success(self.request, 'Tarea creada.')
        return resp


class TaskUpdateView(LoginRequiredMixin, UpdateView):
    model = Task
    form_class = TaskUpdateForm
    template_name = 'core/task_form.html'

    def get_queryset(self):
        return Task.objects.filter(usuario=self.request.user)

    def dispatch(self, request, *args, **kwargs):
        resp = super().dispatch(request, *args, **kwargs)
        tarea = self.get_object()
        if tarea.locked_by and tarea.locked_by != request.user:
            elapsed = (timezone.now() - tarea.locked_at).total_seconds() / 60
            if elapsed < LOCK_TIMEOUT_MINUTES:
                messages.error(request, f'Tarea bloqueada por {tarea.locked_by.username} para edición.')
                return redirect('task_detail', pk=tarea.pk)
        return resp

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['es_editar'] = True
        ctx['fecha_creacion'] = self.object.fecha_creacion
        ctx['archivos'] = self.object.archivos.all()
        return ctx

    def form_valid(self, form):
        old = self.object
        cambios = []
        tracked = {
            'titulo': 'Título', 'comentarios': 'Comentarios', 'estado': 'Estado',
            'prioridad': 'Prioridad', 'departamento': 'Departamento',
            'horas_estimadas': 'Horas est.', 'horas_tomadas': 'Horas tom.',
            'fecha_vencimiento': 'Vencimiento', 'usuario': 'Responsable',
        }
        for field, label in tracked.items():
            old_val = getattr(old, field)
            new_val = form.cleaned_data.get(field)
            if field == 'usuario':
                old_val = old.usuario.username if old.usuario else '-'
                new_val = new_val.username if new_val else '-'
            elif field == 'departamento':
                old_val = str(old.departamento or '-')
                new_val = str(form.cleaned_data.get('departamento') or '-')
            elif field in ('fecha_vencimiento',):
                old_val = str(old_val or '-')
                new_val = str(new_val or '-')
            if old_val != new_val:
                cambios.append(f'{label}: {old_val} → {new_val}')

        form.instance.usuario = form.cleaned_data['usuario']
        if old.usuario != form.instance.usuario:
            crear_notificacion(form.instance.usuario, form.instance, f'Tarea asignada a ti: "{form.instance.titulo}"')
        if old.estado != form.instance.estado:
            if form.instance.estado == 'terminada' and not form.instance.fecha_completada:
                form.instance.fecha_completada = date.today()
                Notification.objects.filter(tarea=form.instance).delete()
                if form.instance.recurrente:
                    self._crear_siguiente_ocurrencia(form.instance)
            elif form.instance.estado != 'terminada':
                form.instance.fecha_completada = None

        resp = super().form_valid(form)
        ets = form.cleaned_data.get('etiquetas')
        if ets is not None: self.object.etiquetas.set(ets)
        for f in self.request.FILES.getlist('archivos'):
            Attachment.objects.create(tarea=self.object, archivo=f, nombre=f.name)
        # Notificar a observadores
        if cambios:
            for watcher in TaskWatcher.objects.filter(tarea=self.object).exclude(usuario=self.request.user):
                crear_notificacion(watcher.usuario, self.object,
                    f'"{self.object.titulo}" actualizada por {self.request.user.username}')
        crear_log(self.object, self.request.user, 'editada', '; '.join(cambios) or 'Editada')
        # Liberar lock
        if self.object.locked_by == self.request.user:
            self.object.locked_by = None
            self.object.locked_at = None
            self.object.save(update_fields=['locked_by', 'locked_at'])
        messages.success(self.request, 'Tarea actualizada.')
        return resp

    def _crear_siguiente_ocurrencia(self, tarea):
        d = {'diaria': timedelta(days=1), 'semanal': timedelta(weeks=1), 'mensual': timedelta(days=30)}.get(tarea.frecuencia)
        if not d: return
        n = Task.objects.create(usuario=tarea.usuario, titulo=tarea.titulo, comentarios=tarea.comentarios,
            prioridad=tarea.prioridad, recurrente=tarea.recurrente, frecuencia=tarea.frecuencia,
            parent=tarea.parent, horas_estimadas=tarea.horas_estimadas,
            creado_por=tarea.creado_por,
            fecha_vencimiento=tarea.fecha_vencimiento + d if tarea.fecha_vencimiento else None,
            recordatorio=tarea.recordatorio + d if tarea.recordatorio else None)
        if tarea.etiquetas.exists(): n.etiquetas.set(tarea.etiquetas.all())


class TaskDeleteView(LoginRequiredMixin, DeleteView):
    model = Task
    template_name = 'core/task_confirm_delete.html'
    success_url = reverse_lazy('task_list')
    context_object_name = 'object'

    def get_queryset(self):
        if Profile.objects.filter(usuario=self.request.user, rol='admin').exists():
            return Task.objects.all()
        return Task.objects.none()

    def delete(self, request, *args, **kwargs):
        tarea = self.get_object()
        _admin_log(request.user, 'eliminar_tarea', f'Tarea "{tarea.titulo}" (id={tarea.id})')
        messages.success(request, 'Tarea eliminada.')
        return super().delete(request, *args, **kwargs)


@login_required
def task_complete(request, pk):
    qs = tareas_visibles(request.user)
    tarea = get_object_or_404(Task, pk=pk)
    if tarea not in qs: raise PermissionDenied
    if request.method == 'POST':
        horas = request.POST.get('horas_tomadas')
        tarea.estado = 'terminada'
        tarea.fecha_completada = date.today()
        if horas: tarea.horas_tomadas = horas
        tarea.save()
        Notification.objects.filter(usuario=request.user, tarea=tarea).delete()
        crear_log(tarea, request.user, 'estado', 'Marcada como terminada')
        if tarea.recurrente:
            d = {'diaria': timedelta(days=1), 'semanal': timedelta(weeks=1), 'mensual': timedelta(days=30)}.get(tarea.frecuencia)
            if d:
                Task.objects.create(usuario=tarea.usuario, titulo=tarea.titulo,
                    comentarios=tarea.comentarios, prioridad=tarea.prioridad,
                    recurrente=tarea.recurrente, frecuencia=tarea.frecuencia,
                    horas_estimadas=tarea.horas_estimadas, creado_por=tarea.creado_por,
                    fecha_vencimiento=tarea.fecha_vencimiento + d if tarea.fecha_vencimiento else None,
                    recordatorio=tarea.recordatorio + d if tarea.recordatorio else None)
                messages.info(request, 'Siguiente ocurrencia creada.')
        messages.success(request, f'"{tarea.titulo}" completada.')
    return redirect('task_detail', pk=pk)


# ─── Kanban ───────────────────────────────────────────────────────

@login_required
def kanban(request):
    qs = tareas_visibles(request.user)
    pendientes = qs.filter(estado='pendiente')
    proceso = qs.filter(estado='proceso')
    terminadas = qs.filter(estado='terminada')
    return render(request, 'core/kanban.html', {
        'pendientes': pendientes, 'proceso': proceso, 'terminadas': terminadas,
    })


@login_required
@require_POST
def kanban_update(request, pk):
    tarea = get_object_or_404(Task, pk=pk)
    if tarea not in tareas_visibles(request.user): return JsonResponse({'error': 'No'}, status=403)
    data = request.POST
    nuevo_estado = data.get('estado')
    if nuevo_estado in ('pendiente', 'proceso', 'terminada'):
        old = tarea.estado
        tarea.estado = nuevo_estado
        if nuevo_estado == 'terminada' and not tarea.fecha_completada:
            tarea.fecha_completada = date.today()
        elif nuevo_estado != 'terminada':
            tarea.fecha_completada = None
        tarea.save()
        crear_log(tarea, request.user, 'estado', f'{old} → {nuevo_estado} (Kanban)')
        return JsonResponse({'ok': True})
    return JsonResponse({'error': 'Estado inválido'}, status=400)


# ─── Task Templates ───────────────────────────────────────────────

@login_required
def template_list(request):
        templates = TaskTemplate.objects.filter(usuario=request.user)
        return render(request, 'core/template_list.html', {'templates': templates})


@login_required
def template_create(request):
    form = TaskTemplateForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        t = form.save(commit=False)
        t.usuario = request.user
        t.save()
        messages.success(request, f'Plantilla "{t.nombre}" creada.')
        return redirect('template_list')
    return render(request, 'core/template_form.html', {'form': form, 'crear': True})


@login_required
def template_edit(request, pk):
    t = get_object_or_404(TaskTemplate, pk=pk, usuario=request.user)
    form = TaskTemplateForm(request.POST or None, instance=t)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Plantilla actualizada.')
        return redirect('template_list')
    return render(request, 'core/template_form.html', {'form': form, 'crear': False, 'tpl': t})


@login_required
def template_delete(request, pk):
    t = get_object_or_404(TaskTemplate, pk=pk, usuario=request.user)
    t.delete()
    messages.success(request, 'Plantilla eliminada.')
    return redirect('template_list')


@login_required
def template_use(request, pk):
    t = get_object_or_404(TaskTemplate, pk=pk, usuario=request.user)
    from django.views.generic.edit import ModelFormMixin
    task = Task.objects.create(
        usuario=request.user, creado_por=request.user,
        titulo=t.titulo, comentarios=t.comentarios,
        prioridad=t.prioridad, horas_estimadas=t.horas_estimadas,
        recurrente=t.recurrente, frecuencia=t.frecuencia,
    )
    crear_notificacion(request.user, task, f'Tarea creada desde plantilla: "{task.titulo}"')
    crear_log(task, request.user, 'creada', f'Creada desde plantilla "{t.nombre}"')
    messages.success(request, f'Tarea creada desde plantilla "{t.nombre}".')
    return redirect('task_detail', pk=task.pk)


# ═══════════════════════════════════════════════════════════════════
#  CALENDAR
# ═══════════════════════════════════════════════════════════════════

@login_required
def calendar_view(request):
    hoy = date.today()
    ano = int(request.GET.get('ano', hoy.year))
    mes = int(request.GET.get('mes', hoy.month))
    if mes < 1: mes, ano = 12, ano - 1
    elif mes > 12: mes, ano = 1, ano + 1
    cal = cal_module.Calendar(firstweekday=0)
    dias_mes = cal.monthdatescalendar(ano, mes)
    pd = date(ano, mes, 1)
    ud = date(ano, mes, cal_module.monthrange(ano, mes)[1])
    qs = Task.objects.filter(usuario=request.user)
    tareas_mes = qs.filter(fecha_vencimiento__gte=pd, fecha_vencimiento__lte=ud)
    tareas_sin_fecha = qs.filter(fecha_vencimiento__isnull=True, estado__in=['pendiente', 'proceso'])
    por_dia = {}
    for t in tareas_mes:
        por_dia.setdefault(t.fecha_vencimiento.day, []).append(t)
    meses = [(i, ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio',
                  'Agosto','Septiembre','Octubre','Noviembre','Diciembre'][i-1]) for i in range(1,13)]
    ctx = {
        'ano': ano, 'mes': mes, 'nombre_mes': dict(meses)[mes],
        'dias_mes': dias_mes, 'tareas_por_dia': por_dia,
        'tareas_sin_fecha': tareas_sin_fecha, 'hoy': hoy,
        'meses': meses, 'anos': list(range(2024, hoy.year + 3)),
        'mes_anterior': mes - 1 if mes > 1 else 12,
        'ano_anterior': ano if mes > 1 else ano - 1,
        'mes_siguiente': mes + 1 if mes < 12 else 1,
        'ano_siguiente': ano if mes < 12 else ano + 1,
    }
    return render(request, 'core/calendar.html', ctx)


# ═══════════════════════════════════════════════════════════════════
#  TAGS
# ═══════════════════════════════════════════════════════════════════

@login_required
def tag_list(request):
    return render(request, 'core/tag_list.html', {'etiquetas': Tag.objects.filter(usuario=request.user)})


@login_required
def tag_create(request):
    if request.method == 'POST':
        n = request.POST.get('nombre', '').strip()
        c = request.POST.get('color', '#58a6ff')
        if n:
            _, created = Tag.objects.get_or_create(usuario=request.user, nombre=n, defaults={'color': c})
            messages.success(request, f'Etiqueta "{n}" {"creada" if created else "ya existe"}.')
        else:
            messages.error(request, 'Nombre obligatorio.')
    return redirect('tag_list')


@login_required
def tag_delete(request, pk):
    t = get_object_or_404(Tag, pk=pk, usuario=request.user)
    t.delete()
    messages.success(request, f'Etiqueta "{t.nombre}" eliminada.')
    return redirect('tag_list')


# ═══════════════════════════════════════════════════════════════════
#  ATTACHMENTS
# ═══════════════════════════════════════════════════════════════════

@login_required
def delete_attachment(request, pk):
    a = get_object_or_404(Attachment, pk=pk, tarea__usuario=request.user)
    tid = a.tarea.pk
    if a.archivo and os.path.isfile(a.archivo.path): os.remove(a.archivo.path)
    a.delete()
    messages.success(request, 'Archivo eliminado.')
    return redirect('task_update', pk=tid)


# ═══════════════════════════════════════════════════════════════════
#  NOTES
# ═══════════════════════════════════════════════════════════════════

class NoteListView(LoginRequiredMixin, ListView):
    model = Note
    template_name = 'core/note_list.html'
    context_object_name = 'notas'
    def get_queryset(self):
        return Note.objects.filter(usuario=self.request.user)


class NoteDetailView(LoginRequiredMixin, DetailView):
    model = Note
    template_name = 'core/note_detail.html'
    context_object_name = 'nota'
    def get_queryset(self):
        return Note.objects.filter(usuario=self.request.user)


class NoteCreateView(LoginRequiredMixin, CreateView):
    model = Note
    form_class = NoteForm
    template_name = 'core/note_form.html'
    def form_valid(self, form):
        form.instance.usuario = self.request.user
        messages.success(self.request, 'Nota creada.')
        return super().form_valid(form)


class NoteUpdateView(LoginRequiredMixin, UpdateView):
    model = Note
    form_class = NoteForm
    template_name = 'core/note_form.html'
    def get_queryset(self):
        return Note.objects.filter(usuario=self.request.user)
    def form_valid(self, form):
        messages.success(self.request, 'Nota actualizada.')
        return super().form_valid(form)


class NoteDeleteView(LoginRequiredMixin, DeleteView):
    model = Note
    template_name = 'core/note_confirm_delete.html'
    success_url = reverse_lazy('note_list')
    context_object_name = 'object'
    def get_queryset(self):
        return Note.objects.filter(usuario=self.request.user)
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Nota eliminada.')
        return super().delete(request, *args, **kwargs)


# ═══════════════════════════════════════════════════════════════════
#  SUMMARY
# ═══════════════════════════════════════════════════════════════════

@login_required
def summary(request):
    ano = int(request.GET.get('ano', datetime.now().year))
    mes = int(request.GET.get('mes', datetime.now().month))
    qs = tareas_visibles(request.user)
    tareas_mes = qs.filter(fecha_creacion__year=ano, fecha_creacion__month=mes)
    completadas = tareas_mes.filter(estado='terminada')
    total_horas = completadas.aggregate(total=Sum('horas_tomadas'))['total'] or 0
    prom = completadas.aggregate(p=Avg('horas_tomadas'))['p'] or 0
    meses = [(i, ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio',
                  'Agosto','Septiembre','Octubre','Noviembre','Diciembre'][i-1]) for i in range(1,13)]

    # Chart: tareas por día del mes
    dias_mes = cal_module.monthrange(ano, mes)[1]
    tareas_por_dia = []
    for d in range(1, dias_mes + 1):
        day_total = tareas_mes.filter(fecha_creacion__day=d).count()
        tareas_por_dia.append(day_total)

    # Chart: distribución de prioridades
    prioridad_labels = ['Urgente', 'Alta', 'Media', 'Baja']
    prioridad_data = [
        tareas_mes.filter(prioridad='urgente').count(),
        tareas_mes.filter(prioridad='alta').count(),
        tareas_mes.filter(prioridad='media').count(),
        tareas_mes.filter(prioridad='baja').count(),
    ]

    # Chart: horas por tarea (top 10)
    top_horas = completadas.filter(horas_tomadas__isnull=False).order_by('-horas_tomadas')[:10]

    return render(request, 'core/summary.html', {
        'ano': ano, 'mes': mes, 'nombre_mes': dict(meses)[mes],
        'total_tareas': tareas_mes.count(),
        'tareas_completadas': completadas,
        'conteo_estados': {
            'pendiente': tareas_mes.filter(estado='pendiente').count(),
            'proceso': tareas_mes.filter(estado='proceso').count(),
            'terminada': completadas.count(),
        },
        'total_horas': total_horas, 'promedio_horas': prom,
        'tareas_con_horas': completadas.filter(horas_tomadas__isnull=False),
        'meses': meses, 'anos': list(range(2024, datetime.now().year + 2)),
        'dias_mes': dias_mes,
        'tareas_por_dia': tareas_por_dia,
        'prioridad_labels': prioridad_labels,
        'prioridad_data': prioridad_data,
        'top_horas': top_horas,
    })


# ═══════════════════════════════════════════════════════════════════
#  NOTIFICATIONS
# ═══════════════════════════════════════════════════════════════════

@login_required
def notification_list(request):
    return render(request, 'core/notifications.html', {
        'notificaciones': Notification.objects.filter(usuario=request.user),
    })


@login_required
def notification_read(request, pk):
    n = get_object_or_404(Notification, pk=pk, usuario=request.user)
    n.leido = True; n.save()
    return redirect('notification_list')


@login_required
def notification_read_all(request):
    Notification.objects.filter(usuario=request.user, leido=False).update(leido=True)
    messages.success(request, 'Todas marcadas como leídas.')
    return redirect('notification_list')


# ═══════════════════════════════════════════════════════════════════
#  NOTIFICATIONS — JSON endpoint for polling
# ═══════════════════════════════════════════════════════════════════

@login_required
def notification_unread_count(request):
    count = Notification.objects.filter(usuario=request.user, leido=False).count()
    return JsonResponse({'count': count})


# ═══════════════════════════════════════════════════════════════════
#  WORKLOAD
# ═══════════════════════════════════════════════════════════════════

@login_required
@admin_mod_required
def workload(request):
    usuarios = User.objects.annotate(
        total=Count('task'),
        pendientes=Count('task', filter=Q(task__estado='pendiente')),
        proceso=Count('task', filter=Q(task__estado='proceso')),
        terminadas=Count('task', filter=Q(task__estado='terminada')),
    ).order_by('username')
    return render(request, 'core/workload.html', {'usuarios': usuarios})


# ═══════════════════════════════════════════════════════════════════
#  EXPORT CSV
# ═══════════════════════════════════════════════════════════════════

@login_required
def export_csv(request):
    qs = tareas_visibles(request.user)
    estado = request.GET.get('estado')
    if estado: qs = qs.filter(estado=estado)
    qs = qs.order_by('-fecha_creacion')
    buf = StringIO()
    w = csv.writer(buf)
    w.writerow(['Título', 'Estado', 'Prioridad', 'Usuario', 'Depto',
                'Horas Est.', 'Horas Tom.', 'Creada', 'Vence', 'Completada'])
    for t in qs:
        w.writerow([t.titulo, t.get_estado_display(), t.get_prioridad_display(),
                    t.usuario.username, t.departamento or '',
                    t.horas_estimadas or '', t.horas_tomadas or '',
                    t.fecha_creacion, t.fecha_vencimiento or '', t.fecha_completada or ''])
    resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = 'attachment; filename="tareas.csv"'
    return resp


# ═══════════════════════════════════════════════════════════════════
#  EXPORT PDF
# ═══════════════════════════════════════════════════════════════════

@login_required
def export_pdf(request):
    if not HAS_REPORTLAB:
        messages.error(request, 'ReportLab no instalado. pip install reportlab')
        return redirect('summary')
    ano = int(request.GET.get('ano', datetime.now().year))
    mes = int(request.GET.get('mes', datetime.now().month))
    qs = tareas_visibles(request.user).filter(fecha_creacion__year=ano, fecha_creacion__month=mes).order_by('fecha_creacion')
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    els = [Paragraph(f'Resumen de Tareas - {mes}/{ano}', styles['Title']), Spacer(1, 12)]
    comp = qs.filter(estado='terminada')
    th = comp.aggregate(t=Sum('horas_tomadas'))['t'] or 0
    els.append(Paragraph(
        f'Total: {qs.count()} | Completadas: {comp.count()} | Pendientes: {qs.filter(estado="pendiente").count()} | Horas: {th:.1f}h',
        styles['Normal']))
    els.append(Spacer(1, 12))
    data = [['Título', 'Estado', 'Usuario', 'Horas', 'Creada']]
    for t in qs:
        data.append([t.titulo[:60], t.get_estado_display(), t.usuario.username, f'{t.horas_tomadas or "-"}h', t.fecha_creacion.strftime('%d/%m/%Y')])
    if len(data) > 1:
        tbl = Table(data, colWidths=[2.5*inch, 1*inch, 1*inch, 0.7*inch, 1*inch])
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#1a73e8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#f5f5f5')]),
        ]))
        els.append(tbl)
    doc.build(els)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="resumen_{ano}_{mes}.pdf"'
    return resp


# ═══════════════════════════════════════════════════════════════════
#  EXPORT XLSX
# ═══════════════════════════════════════════════════════════════════

@login_required
def export_xlsx(request):
    if not HAS_OPENPYXL:
        messages.error(request, 'openpyxl no instalado. pip install openpyxl')
        return redirect('summary')
    qs = tareas_visibles(request.user).order_by('-fecha_creacion')
    estado = request.GET.get('estado')
    if estado: qs = qs.filter(estado=estado)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tareas'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1a73e8')
    headers = ['Título', 'Estado', 'Prioridad', 'Usuario', 'Depto',
               'Horas Est.', 'Horas Tom.', 'Creada', 'Vence', 'Completada']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    for row, t in enumerate(qs, 2):
        ws.cell(row=row, column=1, value=t.titulo)
        ws.cell(row=row, column=2, value=t.get_estado_display())
        ws.cell(row=row, column=3, value=t.get_prioridad_display())
        ws.cell(row=row, column=4, value=t.usuario.username)
        ws.cell(row=row, column=5, value=str(t.departamento or ''))
        ws.cell(row=row, column=6, value=float(t.horas_estimadas) if t.horas_estimadas else '')
        ws.cell(row=row, column=7, value=float(t.horas_tomadas) if t.horas_tomadas else '')
        ws.cell(row=row, column=8, value=t.fecha_creacion.strftime('%Y-%m-%d') if t.fecha_creacion else '')
        ws.cell(row=row, column=9, value=t.fecha_vencimiento.strftime('%Y-%m-%d') if t.fecha_vencimiento else '')
        ws.cell(row=row, column=10, value=t.fecha_completada.strftime('%Y-%m-%d') if t.fecha_completada else '')

    for col in range(1, 11):
        ws.column_dimensions[chr(64 + col)].width = 18

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="tareas.xlsx"'
    wb.save(resp)
    return resp


# ═══════════════════════════════════════════════════════════════════
#  ADMIN PANEL
# ═══════════════════════════════════════════════════════════════════

@login_required
@admin_required
def admin_dashboard(request):
    usuarios = User.objects.annotate(
        total=Count('task'),
        term=Count('task', filter=Q(task__estado='terminada')),
    ).order_by('username')[:20]

    # Chart data
    qs = Task.objects.all()
    task_count_by_user = []
    user_labels = []
    for u in usuarios:
        user_labels.append(u.username)
        task_count_by_user.append(u.total)

    deptos = Department.objects.annotate(total=Count('task'))
    depto_labels = [d.nombre for d in deptos]
    depto_data = [d.total for d in deptos]

    return render(request, 'core/admin/dashboard.html', {
        'total_usuarios': User.objects.count(),
        'total_tareas': qs.count(),
        'tareas_completadas': qs.filter(estado='terminada').count(),
        'tareas_pendientes': qs.filter(estado='pendiente').count(),
        'tareas_proceso': qs.filter(estado='proceso').count(),
        'total_departamentos': Department.objects.count(),
        'resumen_usuarios': usuarios,
        'chart_tareas_por_usuario_labels': user_labels,
        'chart_tareas_por_usuario_data': task_count_by_user,
        'chart_depto_labels': depto_labels,
        'chart_depto_data': depto_data,
        'login_logs': LoginLog.objects.all()[:10],
        'admin_logs': AdminLog.objects.all()[:15],
    })


@login_required
@admin_required
def admin_users(request):
    return render(request, 'core/admin/user_list.html', {
        'usuarios': User.objects.all().order_by('username'),
    })


@login_required
@admin_required
def admin_user_create(request):
    form = AdminUserCreateForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        if User.objects.filter(username=form.cleaned_data['username']).exists():
            messages.error(request, 'El usuario ya existe.')
        else:
            u = User.objects.create_user(username=form.cleaned_data['username'],
                password=form.cleaned_data['password'], email=form.cleaned_data.get('email',''),
                first_name=form.cleaned_data.get('first_name',''), last_name=form.cleaned_data.get('last_name',''))
            p, _ = Profile.objects.get_or_create(usuario=u)
            p.rol = form.cleaned_data['rol']; p.save()
            p.departamentos.set(form.cleaned_data.get('departamentos', []))
            PasswordHistory.objects.create(usuario=u, password=u.password)
            _admin_log(request.user, 'crear_usuario', f'Usuario "{u.username}" rol={p.rol}')
            messages.success(request, f'Usuario "{u.username}" creado.')
            return redirect('admin_users')
    return render(request, 'core/admin/user_form.html', {'form': form, 'crear': True})


@login_required
@admin_required
def admin_user_edit(request, pk):
    u = get_object_or_404(User, pk=pk)
    p, _ = Profile.objects.get_or_create(usuario=u)
    form = AdminUserEditForm(request.POST or None, initial={
        'email': u.email, 'first_name': u.first_name, 'last_name': u.last_name,
        'is_active': u.is_active, 'rol': p.rol, 'departamentos': p.departamentos.all(),
    })
    if request.method == 'POST' and form.is_valid():
        for fld in ('email', 'first_name', 'last_name'):
            setattr(u, fld, form.cleaned_data[fld])
        u.is_active = form.cleaned_data['is_active']
        if form.cleaned_data['nueva_password']:
            u.set_password(form.cleaned_data['nueva_password'])
        u.save()
        if form.cleaned_data['nueva_password']:
            PasswordHistory.objects.create(usuario=u, password=u.password)
        p.rol = form.cleaned_data['rol']; p.save()
        p.departamentos.set(form.cleaned_data.get('departamentos', []))
        cambios = []
        for fld in ('email', 'first_name', 'last_name', 'is_active'):
            old = getattr(u, fld) if fld != 'is_active' else u.is_active
            new = form.cleaned_data[fld]
            if old != new:
                cambios.append(f'{fld}: {old} → {new}')
        if form.cleaned_data['nueva_password']:
            cambios.append('password cambiada')
        if p.rol != Profile.objects.get(usuario=u).rol:
            cambios.append(f'rol: → {p.rol}')
        _admin_log(request.user, 'editar_usuario', f'Usuario "{u.username}": {"; ".join(cambios)}')
        messages.success(request, f'Usuario "{u.username}" actualizado.')
        return redirect('admin_users')
    return render(request, 'core/admin/user_form.html', {'form': form, 'crear': False, 'edit_user': u})


@login_required
@admin_mod_required
def admin_tasks(request):
    qs = Task.objects.all().select_related('usuario', 'departamento')
    for k in ('estado', 'usuario', 'departamento'):
        v = request.GET.get(k)
        if v:
            qs = qs.filter(**{k: v})
    paginator = Paginator(qs, 30)
    page = request.GET.get('page', 1)
    tareas_page = paginator.get_page(page)
    return render(request, 'core/admin/task_list.html', {
        'tareas': tareas_page, 'usuarios': User.objects.all(), 'departamentos': Department.objects.all(),
        'estado_actual': request.GET.get('estado',''),
        'usuario_actual': request.GET.get('usuario',''),
        'depto_actual': request.GET.get('departamento',''),
    })


@login_required
@admin_required
def admin_departments(request):
    return render(request, 'core/admin/department_list.html', {'departamentos': Department.objects.all()})


@login_required
@admin_required
def admin_department_create(request):
    f = AdminDepartmentForm(request.POST or None)
    if request.method == 'POST' and f.is_valid(): f.save(); messages.success(request, 'Departamento creado.'); return redirect('admin_departments')
    return render(request, 'core/admin/department_form.html', {'form': f, 'crear': True})


@login_required
@admin_required
def admin_department_edit(request, pk):
    d = get_object_or_404(Department, pk=pk)
    f = AdminDepartmentForm(request.POST or None, instance=d)
    if request.method == 'POST' and f.is_valid(): f.save(); messages.success(request, 'Departamento actualizado.'); return redirect('admin_departments')
    return render(request, 'core/admin/department_form.html', {'form': f, 'crear': False, 'depto': d})


BACKUP_DIR = settings.BASE_DIR / 'backups'

@login_required
@admin_required
def admin_backups(request):
    backup_dir = BACKUP_DIR
    backup_dir.mkdir(parents=True, exist_ok=True)

    if request.method == 'POST':
        accion = request.POST.get('accion')
        if accion == 'crear':
            call_command('backup_tareas')
            _admin_log(request.user, 'backup', 'Copia de seguridad creada desde admin')
            messages.success(request, 'Copia de seguridad creada.')
        elif accion == 'eliminar':
            filename = request.POST.get('filename', '')
            if filename:
                filepath = backup_dir / filename
                if filepath.exists() and filepath.is_file():
                    filepath.unlink()
                    messages.success(request, f'Backup "{filename}" eliminado.')
        return redirect('admin_backups')

    if 'descargar' in request.GET:
        filename = request.GET['descargar']
        filepath = backup_dir / filename
        if filepath.exists() and filepath.is_file():
            with open(filepath, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/octet-stream')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
        messages.error(request, 'Archivo no encontrado.')
        return redirect('admin_backups')

    backups = []
    for f in sorted(backup_dir.iterdir(), key=lambda p: p.stat().st_mtime, reverse=True):
        if f.suffix == '.sqlite3':
            stat = f.stat()
            size_kb = stat.st_size / 1024
            backups.append({
                'filename': f.name,
                'size': f'{size_kb:.1f} KB' if size_kb < 1024 else f'{size_kb/1024:.1f} MB',
                'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
            })

    return render(request, 'core/admin/backup_list.html', {'backups': backups})


# ═══════════════════════════════════════════════════════════════════
#  PER-USER KANBAN BOARD
# ═══════════════════════════════════════════════════════════════════

@login_required
@admin_mod_required
def admin_board(request):
    usuarios = User.objects.all().order_by('username')
    board_data = []
    for u in usuarios:
        qs = Task.objects.filter(usuario=u)
        board_data.append({
            'usuario': u,
            'pendientes': qs.filter(estado='pendiente'),
            'proceso': qs.filter(estado='proceso'),
            'terminadas': qs.filter(estado='terminada'),
            'total': qs.count(),
        })
    return render(request, 'core/admin/board.html', {'board_data': board_data})


# ═══════════════════════════════════════════════════════════════════
#  SUBTASK REORDER
# ═══════════════════════════════════════════════════════════════════

@login_required
@require_POST
def subtask_reorder(request, pk):
    tarea = get_object_or_404(Task, pk=pk)
    if tarea not in tareas_visibles(request.user):
        return JsonResponse({'error': 'No'}, status=403)
    ids = request.POST.getlist('subtask_ids[]')
    if not ids:
        return JsonResponse({'error': 'Sin datos'}, status=400)
    for i, sid in enumerate(ids):
        subtask = get_object_or_404(Task, pk=sid, parent=tarea)
        subtask.orden = i
        subtask.save(update_fields=['orden'])
    return JsonResponse({'ok': True})


# ═══════════════════════════════════════════════════════════════════
#  SHARE TASK
# ═══════════════════════════════════════════════════════════════════

@login_required
@require_POST
def generate_share_token(request, pk):
    tarea = get_object_or_404(Task, pk=pk)
    if tarea not in tareas_visibles(request.user):
        return JsonResponse({'error': 'No'}, status=403)
    if not tarea.share_token:
        tarea.share_token = uuid.uuid4()
        tarea.save(update_fields=['share_token'])
    return JsonResponse({'token': str(tarea.share_token)})


def shared_task(request, token):
    tarea = get_object_or_404(Task, share_token=token)
    return render(request, 'core/shared_task.html', {'tarea': tarea})


# ─── Watch / Unwatch ──────────────────────────────────────────

@login_required
@require_POST
def watch_task(request, pk):
    tarea = get_object_or_404(Task, pk=pk)
    if tarea not in tareas_visibles(request.user):
        return JsonResponse({'error': 'No'}, status=403)
    _, created = TaskWatcher.objects.get_or_create(usuario=request.user, tarea=tarea)
    return JsonResponse({'watching': True, 'created': created})


@login_required
@require_POST
def unwatch_task(request, pk):
    tarea = get_object_or_404(Task, pk=pk)
    TaskWatcher.objects.filter(usuario=request.user, tarea=tarea).delete()
    return JsonResponse({'watching': False})


# ─── Task Locking ─────────────────────────────────────────────

LOCK_TIMEOUT_MINUTES = 10


@login_required
@require_POST
def task_lock(request, pk):
    tarea = get_object_or_404(Task, pk=pk)
    if tarea not in tareas_visibles(request.user):
        return JsonResponse({'error': 'No'}, status=403)
    if tarea.locked_by and tarea.locked_by != request.user:
        elapsed = (timezone.now() - tarea.locked_at).total_seconds() / 60
        if elapsed < LOCK_TIMEOUT_MINUTES:
            return JsonResponse({'locked': True, 'by': tarea.locked_by.username})
    tarea.locked_by = request.user
    tarea.locked_at = timezone.now()
    tarea.save(update_fields=['locked_by', 'locked_at'])
    return JsonResponse({'locked': True, 'by': request.user.username})


@login_required
@require_POST
def task_unlock(request, pk):
    tarea = get_object_or_404(Task, pk=pk)
    if tarea.locked_by == request.user:
        tarea.locked_by = None
        tarea.locked_at = None
        tarea.save(update_fields=['locked_by', 'locked_at'])
    return JsonResponse({'locked': False})


@login_required
@require_POST
def task_change_status(request, pk):
    tarea = get_object_or_404(Task, pk=pk)
    if tarea not in tareas_visibles(request.user):
        return JsonResponse({'error': 'No'}, status=403)
    estado = request.POST.get('estado')
    if estado not in dict(Task.ESTADOS):
        return JsonResponse({'error': 'Estado inválido'}, status=400)
    tarea.estado = estado
    if estado == 'terminada':
        tarea.fecha_completada = date.today()
        Notification.objects.filter(usuario=request.user, tarea=tarea).delete()
        if tarea.recurrente:
            d = {'diaria': timedelta(days=1), 'semanal': timedelta(weeks=1), 'mensual': timedelta(days=30)}.get(tarea.frecuencia)
            if d:
                Task.objects.create(usuario=tarea.usuario, titulo=tarea.titulo,
                    comentarios=tarea.comentarios, prioridad=tarea.prioridad,
                    recurrente=tarea.recurrente, frecuencia=tarea.frecuencia,
                    horas_estimadas=tarea.horas_estimadas, creado_por=tarea.creado_por,
                    fecha_vencimiento=tarea.fecha_vencimiento + d if tarea.fecha_vencimiento else None,
                    recordatorio=tarea.recordatorio + d if tarea.recordatorio else None)
    tarea.save(update_fields=['estado', 'fecha_completada'])
    crear_log(tarea, request.user, 'estado', f'Estado cambiado a {tarea.get_estado_display()}')
    # Notificar observadores
    for watcher in TaskWatcher.objects.filter(tarea=tarea).exclude(usuario=request.user):
        crear_notificacion(watcher.usuario, tarea,
            f'Estado de "{tarea.titulo}" cambiado a {tarea.get_estado_display()} por {request.user.username}')
    return JsonResponse({'ok': True, 'estado': estado, 'display': tarea.get_estado_display()})


@login_required
def profile(request):
    u = request.user
    p, _ = Profile.objects.get_or_create(usuario=u)
    form = ProfileForm(request.POST or None, request.FILES or None, instance=u, initial={
        'username': u.username,
        'first_name': u.first_name,
        'last_name': u.last_name,
        'email': u.email,
    })
    if request.method == 'POST' and form.is_valid():
        nuevo_username = form.cleaned_data['username']
        if nuevo_username != u.username:
            if User.objects.filter(username=nuevo_username).exclude(pk=u.pk).exists():
                messages.error(request, 'El nombre de usuario ya está en uso.')
                return render(request, 'core/profile.html', {'form': form, 'profile': p})
            u.username = nuevo_username
        u.first_name = form.cleaned_data['first_name']
        u.last_name = form.cleaned_data['last_name']
        u.email = form.cleaned_data['email']
        password = form.cleaned_data.get('password1')
        if password:
            u.set_password(password)
        u.save()
        if password:
            PasswordHistory.objects.create(usuario=u, password=u.password)

        avatar_file = form.cleaned_data.get('avatar')
        icon = request.POST.get('avatar_icon', '')
        if form.cleaned_data.get('eliminar_avatar'):
            if p.avatar:
                p.avatar.delete(save=False)
            p.avatar = None
        elif avatar_file:
            if p.avatar:
                p.avatar.delete(save=False)
            p.avatar = avatar_file

        p.avatar_icon = icon
        p.save()

        messages.success(request, 'Perfil actualizado.')
        if password:
            from django.contrib.auth import update_session_auth_hash
            update_session_auth_hash(request, u)
            messages.success(request, 'Contraseña cambiada.')
        return redirect('profile')
    return render(request, 'core/profile.html', {'form': form, 'profile': p})
